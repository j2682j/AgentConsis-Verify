from __future__ import annotations

from collections import deque
from pathlib import Path
import re
from typing import Any

from ..base import HandlerInput, HandlerMatch, HandlerResult
from ..contracts import default_outputs, input_field, io_contract


class ColorGridHamiltonianRouterHandler:
    """判斷同色正交網格是否存在走訪每格一次的封閉路徑。"""

    name = "color_grid_hamiltonian"
    handler_role = "grid_hamiltonian_cycle"
    uses_specialized_attachment_parser = True
    capability_description = (
        "Determine whether all cells of one color in a spreadsheet grid admit a "
        "Hamiltonian cycle using orthogonal adjacency."
    )
    supported_attachment_types = {".xlsx", ".xlsm"}
    supported_task_roles = {"grid_hamiltonian_cycle", "hamiltonian_cycle"}
    supported_answer_roles = {"boolean", "yes_no"}
    input_schema = io_contract(
        name,
        [
            input_field("file_path", "str", True, "Spreadsheet path.", "attachment"),
            input_field("target_color", "str", True, "Requested cell color.", "question"),
        ],
        default_outputs(),
        supported_attachment_types=supported_attachment_types,
    )
    output_schema = input_schema

    _COLOR_RGB = {
        "green": "FF00FF00",
        "blue": "FF4A86E8",
        "red": "FFFF0000",
        "yellow": "FFFFFF00",
        "orange": "FFFF9900",
        "purple": "FF9900FF",
    }

    def match_input(self, handler_input: HandlerInput) -> HandlerMatch:
        path = self._file_path(handler_input)
        question = handler_input.question.casefold()
        color = self._target_color(question)
        operation = bool(
            re.search(r"\b(?:every|each|all)\b", question)
            and re.search(r"\b(?:return|back)\b", question)
            and re.search(r"\b(?:starting|start)\b", question)
        )
        ready = bool(
            path
            and Path(path).suffix.casefold() in self.supported_attachment_types
            and color
            and operation
        )
        return HandlerMatch(
            handler_name=self.name,
            matched=ready,
            confidence=0.99 if ready else 0.0,
            reason="colored_grid_closed_visit_request",
            handler_role=self.handler_role,
            missing_inputs=[] if ready else ["spreadsheet_color_or_closed_visit_requirement"],
        )

    def build_input(self, handler_input: HandlerInput) -> dict[str, Any]:
        return {
            "file_path": self._file_path(handler_input),
            "target_color": self._target_color(handler_input.question.casefold()),
        }

    def run(self, inputs: dict[str, Any]) -> HandlerResult:
        from openpyxl import load_workbook

        file_path = str(inputs.get("file_path") or "").strip()
        color_name = str(inputs.get("target_color") or "").strip().casefold()
        rgb = self._COLOR_RGB.get(color_name, color_name.upper())
        if not file_path or not Path(file_path).is_file() or not rgb:
            return HandlerResult.missing(
                handler_name=self.name,
                missing_inputs=["file_path", "target_color"],
            )
        workbook = load_workbook(file_path, data_only=True, read_only=False)
        sheet = workbook.active
        nodes = {
            (cell.row, cell.column)
            for row in sheet.iter_rows()
            for cell in row
            if self._cell_rgb(cell) == rgb
        }
        if not nodes:
            return HandlerResult.missing(
                handler_name=self.name,
                missing_inputs=["target_color_cells"],
            )
        adjacency = {
            node: sorted(
                neighbor
                for neighbor in self._neighbors(node)
                if neighbor in nodes
            )
            for node in nodes
        }
        components = self._components(adjacency)
        low_degree = sorted(node for node, edges in adjacency.items() if len(edges) < 2)
        partitions = (
            sum((row + column) % 2 == 0 for row, column in nodes),
            sum((row + column) % 2 == 1 for row, column in nodes),
        )
        witness = ""
        answer = ""
        cycle: list[tuple[int, int]] = []
        if len(components) != 1:
            answer = "no"
            witness = "grid_is_disconnected"
        elif low_degree:
            answer = "no"
            witness = "a_cycle_vertex_has_degree_below_two"
        elif partitions[0] != partitions[1]:
            answer = "no"
            witness = "bipartite_partition_sizes_are_unequal"
        elif len(nodes) <= 28:
            cycle = self._find_cycle(adjacency)
            answer = "yes" if cycle else "no"
            witness = "explicit_cycle" if cycle else "exhaustive_search_found_no_cycle"
        else:
            return HandlerResult.missing(
                handler_name=self.name,
                missing_inputs=["large_grid_exact_cycle_solver"],
                structured_result={
                    "node_count": len(nodes),
                    "edge_count": sum(map(len, adjacency.values())) // 2,
                    "connected_components": len(components),
                    "bipartite_partitions": list(partitions),
                },
                next_action_hint="Use an exact CP-SAT Hamiltonian-cycle solver for this large inconclusive grid.",
            )
        coordinates = [self._coordinate(node) for node in sorted(nodes)]
        evidence = (
            f"The {color_name} grid has {len(nodes)} vertices and "
            f"{sum(map(len, adjacency.values())) // 2} orthogonal edges. "
            f"Hamiltonian-cycle result: {answer}. Witness: {witness}."
        )
        return HandlerResult(
            handler_name=self.name,
            status="ok",
            answer=answer,
            evidence_text=evidence,
            confidence=1.0,
            output_type="final_answer",
            semantic_role="hamiltonian_cycle_exists",
            supporting_inputs=[file_path, color_name, *coordinates],
            structured_result={
                "task_type": "grid_hamiltonian_cycle",
                "operation": "hamiltonian_cycle",
                "target_color": color_name,
                "node_count": len(nodes),
                "edge_count": sum(map(len, adjacency.values())) // 2,
                "connected_components": len(components),
                "degree_by_node": {
                    self._coordinate(node): len(adjacency[node]) for node in sorted(nodes)
                },
                "bipartite_partitions": list(partitions),
                "failure_witness": witness if answer == "no" else "",
                "cycle_path": [self._coordinate(node) for node in cycle],
                "input_provenance": {
                    "source": "specialized_attachment_input",
                    "file_path": file_path,
                    "parse_status": "success",
                },
            },
        )

    @staticmethod
    def _cell_rgb(cell: Any) -> str:
        value = getattr(getattr(cell.fill, "fgColor", None), "rgb", "")
        return str(value or "").upper()

    @classmethod
    def _target_color(cls, question: str) -> str:
        return next((name for name in cls._COLOR_RGB if re.search(rf"\b{name}\b", question)), "")

    @staticmethod
    def _file_path(handler_input: HandlerInput) -> str:
        adapted = handler_input.adapted_inputs()
        attachment = handler_input.attachment if isinstance(handler_input.attachment, dict) else {}
        return str(
            adapted.get("file_path")
            or attachment.get("file_path")
            or attachment.get("path")
            or ""
        ).strip()

    @staticmethod
    def _neighbors(node: tuple[int, int]) -> tuple[tuple[int, int], ...]:
        row, column = node
        return ((row - 1, column), (row + 1, column), (row, column - 1), (row, column + 1))

    @staticmethod
    def _components(adjacency: dict[tuple[int, int], list[tuple[int, int]]]) -> list[set[tuple[int, int]]]:
        remaining = set(adjacency)
        output: list[set[tuple[int, int]]] = []
        while remaining:
            start = next(iter(remaining))
            seen = {start}
            queue = deque([start])
            while queue:
                node = queue.popleft()
                for neighbor in adjacency[node]:
                    if neighbor not in seen:
                        seen.add(neighbor)
                        queue.append(neighbor)
            remaining -= seen
            output.append(seen)
        return output

    @staticmethod
    def _find_cycle(adjacency: dict[tuple[int, int], list[tuple[int, int]]]) -> list[tuple[int, int]]:
        start = min(adjacency, key=lambda node: len(adjacency[node]))
        total = len(adjacency)
        path = [start]
        visited = {start}

        def visit(node: tuple[int, int]) -> bool:
            if len(path) == total:
                return start in adjacency[node]
            options = sorted(
                (item for item in adjacency[node] if item not in visited),
                key=lambda item: sum(neighbor not in visited for neighbor in adjacency[item]),
            )
            for neighbor in options:
                visited.add(neighbor)
                path.append(neighbor)
                if visit(neighbor):
                    return True
                path.pop()
                visited.remove(neighbor)
            return False

        return [*path, start] if visit(start) else []

    @staticmethod
    def _coordinate(node: tuple[int, int]) -> str:
        from openpyxl.utils import get_column_letter

        return f"{get_column_letter(node[1])}{node[0]}"


__all__ = ["ColorGridHamiltonianRouterHandler"]
