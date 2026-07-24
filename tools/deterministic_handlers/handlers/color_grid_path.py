from __future__ import annotations

from collections import deque
from pathlib import Path
import re
from typing import Any

from ..base import HandlerInput, HandlerMatch, HandlerResult
from ..contracts import default_outputs, input_field, io_contract


_ORDINALS = {
    "first": 1, "second": 2, "third": 3, "fourth": 4, "fifth": 5,
    "sixth": 6, "seventh": 7, "eighth": 8, "ninth": 9, "tenth": 10,
    "eleventh": 11, "twelfth": 12, "thirteenth": 13, "fourteenth": 14,
    "fifteenth": 15, "sixteenth": 16, "seventeenth": 17, "eighteenth": 18,
    "nineteenth": 19, "twentieth": 20,
}
_SMALL_NUMBERS = {
    "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
    "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
}
_COLOR_PROTOTYPES = {
    "blue": (0, 0, 255),
    "red": (255, 0, 0),
    "green": (0, 176, 80),
    "yellow": (255, 255, 0),
    "orange": (255, 153, 0),
    "purple": (153, 0, 255),
    "pink": (244, 120, 167),
    "black": (0, 0, 0),
    "white": (255, 255, 255),
}


class ColorGridPathRouterHandler:
    """走訪試算表色格迷宮並回報第 N 回合落點格子的填色代碼。"""

    name = "color_grid_path"
    handler_role = "grid_turn_navigation"
    uses_specialized_attachment_parser = True
    capability_description = (
        "Walk a spreadsheet color-grid maze from START toward END with a fixed "
        "number of cells per turn while avoiding a named color, and report the "
        "fill color hex of the cell reached on a requested turn."
    )
    supported_attachment_types = {".xlsx", ".xlsm"}
    supported_task_roles = {"grid_turn_navigation", "grid_path_color"}
    supported_answer_roles = {"color_hex", "string"}
    input_schema = io_contract(
        name,
        [
            input_field("file_path", "str", True, "Spreadsheet path.", "attachment"),
            input_field("avoid_color", "str", True, "Color name that may not be stepped on.", "question"),
            input_field("turn_number", "int", True, "Turn whose landing cell is requested.", "question"),
            input_field("cells_per_turn", "int", True, "Cells moved on each turn.", "question"),
        ],
        default_outputs(),
        supported_attachment_types=supported_attachment_types,
    )
    output_schema = input_schema

    def match_input(self, handler_input: HandlerInput) -> HandlerMatch:
        path = self._file_path(handler_input)
        question = handler_input.question.casefold()
        ready = bool(
            path
            and Path(path).suffix.casefold() in self.supported_attachment_types
            and self._avoid_color(question)
            and self._turn_number(question)
            and re.search(r"\bhex\b|\bcolou?r\b", question)
            and re.search(r"\bstart\b", question)
            and re.search(r"\bturn\b", question)
        )
        return HandlerMatch(
            handler_name=self.name,
            matched=ready,
            confidence=0.99 if ready else 0.0,
            reason="grid_turn_landing_color_request",
            handler_role=self.handler_role,
            missing_inputs=[] if ready else ["grid_turn_navigation_inputs"],
        )

    def build_input(self, handler_input: HandlerInput) -> dict[str, Any]:
        question = handler_input.question.casefold()
        return {
            "file_path": self._file_path(handler_input),
            "avoid_color": self._avoid_color(question),
            "turn_number": self._turn_number(question),
            "cells_per_turn": self._cells_per_turn(question),
        }

    def run(self, inputs: dict[str, Any]) -> HandlerResult:
        from openpyxl import load_workbook

        file_path = str(inputs.get("file_path") or "").strip()
        avoid_color = str(inputs.get("avoid_color") or "").strip().casefold()
        turn_number = int(inputs.get("turn_number") or 0)
        cells_per_turn = max(1, int(inputs.get("cells_per_turn") or 1))
        if not file_path or not Path(file_path).is_file() or not avoid_color or turn_number <= 0:
            return HandlerResult.missing(
                handler_name=self.name,
                missing_inputs=["file_path", "avoid_color", "turn_number"],
            )
        workbook = load_workbook(file_path, data_only=True, read_only=False)
        sheet = workbook.active
        grid: dict[tuple[int, int], str] = {}
        start = end = None
        for row in sheet.iter_rows():
            for cell in row:
                grid[(cell.row, cell.column)] = self._cell_rgb(cell)
                value = str(cell.value or "").strip().casefold()
                if value == "start":
                    start = (cell.row, cell.column)
                elif value == "end":
                    end = (cell.row, cell.column)
        if start is None or end is None:
            return HandlerResult.missing(
                handler_name=self.name,
                missing_inputs=["start_and_end_cells"],
            )
        avoid_rgb = self._resolve_color(avoid_color, grid.values())
        if not avoid_rgb:
            return HandlerResult.missing(
                handler_name=self.name,
                missing_inputs=["avoid_color_cells"],
            )

        walkable = {
            node for node, rgb in grid.items() if rgb != avoid_rgb
        }
        forward = self._distances(start, walkable)
        backward = self._distances(end, walkable)
        if end not in forward:
            return HandlerResult.missing(
                handler_name=self.name,
                missing_inputs=["path_from_start_to_end"],
            )
        total = forward[end]
        target_steps = turn_number * cells_per_turn
        if target_steps > total:
            return HandlerResult.missing(
                handler_name=self.name,
                missing_inputs=["turn_number_within_path_length"],
                structured_result={"shortest_path_cells": total},
            )
        # Every cell on a shortest path at exactly target_steps steps: the
        # answer is deterministic only when all of them share one fill color.
        landing_cells = sorted(
            node
            for node in walkable
            if forward.get(node) == target_steps
            and backward.get(node, total + 1) + target_steps == total
        )
        landing_colors = {self._hex6(grid[node]) for node in landing_cells}
        if len(landing_colors) != 1:
            return HandlerResult.missing(
                handler_name=self.name,
                missing_inputs=["unique_landing_cell_color"],
                structured_result={
                    "landing_cells": [self._coordinate(node) for node in landing_cells],
                    "landing_colors": sorted(landing_colors),
                },
            )
        answer = landing_colors.pop()
        landing = landing_cells[0]
        evidence = (
            f"Shortest {cells_per_turn}-cell-per-turn walk from START "
            f"{self._coordinate(start)} toward END {self._coordinate(end)} avoiding "
            f"{avoid_color} ({avoid_rgb}) lands on {self._coordinate(landing)} after "
            f"turn {turn_number}; its fill color is {answer}."
        )
        return HandlerResult(
            handler_name=self.name,
            status="ok",
            answer=answer,
            evidence_text=evidence,
            confidence=1.0,
            output_type="final_answer",
            semantic_role="grid_turn_landing_color",
            supporting_inputs=[
                file_path,
                avoid_color,
                str(turn_number),
                str(cells_per_turn),
                self._coordinate(landing),
            ],
            structured_result={
                "task_type": "grid_turn_navigation",
                "operation": "grid_turn_navigation",
                "avoid_color": avoid_color,
                "avoid_rgb": avoid_rgb,
                "turn_number": turn_number,
                "cells_per_turn": cells_per_turn,
                "landing_cell": self._coordinate(landing),
                "landing_candidates": [self._coordinate(node) for node in landing_cells],
                "shortest_path_cells": total,
                "input_provenance": {
                    "source": "specialized_attachment_input",
                    "file_path": file_path,
                    "parse_status": "success",
                },
            },
        )

    @staticmethod
    def _distances(
        origin: tuple[int, int],
        walkable: set[tuple[int, int]],
    ) -> dict[tuple[int, int], int]:
        distances = {origin: 0}
        queue = deque([origin])
        while queue:
            row, column = queue.popleft()
            for neighbor in (
                (row - 1, column),
                (row + 1, column),
                (row, column - 1),
                (row, column + 1),
            ):
                if neighbor in walkable and neighbor not in distances:
                    distances[neighbor] = distances[(row, column)] + 1
                    queue.append(neighbor)
        return distances

    @classmethod
    def _resolve_color(cls, color_name: str, grid_colors: Any) -> str:
        prototype = cls._COLOR_PROTOTYPES_SAFE(color_name)
        if prototype is None:
            return ""
        best_rgb = ""
        best_distance = None
        for rgb in {str(value or "") for value in grid_colors}:
            parsed = cls._parse_rgb(rgb)
            if parsed is None:
                continue
            distance = sum((a - b) ** 2 for a, b in zip(parsed, prototype))
            if best_distance is None or distance < best_distance:
                best_distance = distance
                best_rgb = rgb
        return best_rgb

    @staticmethod
    def _COLOR_PROTOTYPES_SAFE(color_name: str) -> tuple[int, int, int] | None:
        return _COLOR_PROTOTYPES.get(str(color_name or "").strip().casefold())

    @staticmethod
    def _parse_rgb(value: str) -> tuple[int, int, int] | None:
        text = str(value or "").strip().upper()
        if len(text) == 8:
            text = text[2:]
        if len(text) != 6 or not re.fullmatch(r"[0-9A-F]{6}", text):
            return None
        return tuple(int(text[index:index + 2], 16) for index in (0, 2, 4))

    @staticmethod
    def _hex6(value: str) -> str:
        text = str(value or "").strip().upper()
        if len(text) == 8:
            text = text[2:]
        return text

    @staticmethod
    def _cell_rgb(cell: Any) -> str:
        value = getattr(getattr(cell.fill, "fgColor", None), "rgb", "")
        return str(value or "").upper()

    @classmethod
    def _avoid_color(cls, question: str) -> str:
        match = re.search(
            r"avoid[^.]*?\b(" + "|".join(_COLOR_PROTOTYPES) + r")\b"
            r"|\bnot\b[^.]*?\b(" + "|".join(_COLOR_PROTOTYPES) + r")\b[^.]*?\b(?:cells?|background)\b",
            question,
        )
        if not match:
            return ""
        return next(group for group in match.groups() if group)

    @staticmethod
    def _turn_number(question: str) -> int:
        match = re.search(r"\b(\d+)(?:st|nd|rd|th)\s+turn\b", question)
        if match:
            return int(match.group(1))
        match = re.search(
            r"\b(" + "|".join(_ORDINALS) + r")\s+turn\b",
            question,
        )
        if match:
            return _ORDINALS[match.group(1)]
        return 0

    @staticmethod
    def _cells_per_turn(question: str) -> int:
        match = re.search(
            r"\bmove\s+(\d+|" + "|".join(_SMALL_NUMBERS) + r")\s+cells?\b",
            question,
        )
        if not match:
            return 1
        token = match.group(1)
        return int(token) if token.isdigit() else _SMALL_NUMBERS[token]

    @staticmethod
    def _file_path(handler_input: HandlerInput) -> str:
        adapted = handler_input.adapted_inputs()
        attachment = handler_input.attachment if isinstance(handler_input.attachment, dict) else {}
        return str(
            adapted.get("file_path")
            or attachment.get("file_path")
            or attachment.get("path")
            or ""
        ).strip()

    @staticmethod
    def _coordinate(node: tuple[int, int]) -> str:
        from openpyxl.utils import get_column_letter

        return f"{get_column_letter(node[1])}{node[0]}"


__all__ = ["ColorGridPathRouterHandler"]
